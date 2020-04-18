from urllib import request
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Alignment, Font, numbers
import smtplib
from email.message import EmailMessage



# Входные данные ----------------------------------------------------------
web_page1 = 'https://yandex.ru/news/quotes/2002.html'       # Доллар
web_page2 = 'https://yandex.ru/news/quotes/2000.html'       # Евро
RESULTFILE = 'PythonResult.xlsx'
SMTP_SERVER = "smtp.mail.ru"
SMTP_PORT = 465
SENDER_LOGIN = "a.a.test@mail.ru"
SENDER_PASSWORD = "uipathTest"
RECIEVER_EMAIL = "antohnio@mail.ru"
SUBJECT = "Результат теста по скраппингу курсов валют - А.А. Макаров, Python"
BEGIN_MESSAGE = "Это результат теста с курсами валют на Питоне. В файле "


# Выгружаем данные с Яндекса -----------------------------------------------
def cut(source, begin, end, startindex=0):
    Ind1 = source[startindex:].index(begin)+len(begin)
    Ind2 = source[(startindex + Ind1):].index(end)
    finalindex = startindex + Ind1 + Ind2
    result = source[(startindex+Ind1):(startindex+Ind1+Ind2)]
    return result, finalindex


def scrapper(web_site):
    req = request.Request(web_site)
    response = request.urlopen(req)
    web_page = str(response.read())
    table = cut(web_page, '</th></tr><tr class="quote__day quote', '</table>')[0]
    index=0
    date = ['Дата']
    value = ['Курс']
    change = ['Изменение']
    try:
        while table:
            currentdate = cut(table, '<td class="quote__date">', '</t', index)
            date.append(currentdate[0])
            index = currentdate[1]
            curerentvalue = cut(table, '<td class="quote__value"><span class="quote__sgn"></span>', '</t', index)
            value.append(curerentvalue[0])
            index = curerentvalue[1]
            currentsign = cut(table, '<span class="quote__sgn">', '</span>', index)
            sign = currentsign[0]
            index = currentsign[1]
            currentchange = cut(table, '</span>', '</t', index)
            change.append(sign+currentchange[0])
            index = currentchange[1]
    except ValueError:
        pass
    return date, value, change


D = scrapper(web_page1)
E = scrapper(web_page2)


# Работаем с файлом Эксель -----------------------------------------------
wb = Workbook()
dest_filename = RESULTFILE
ws1 = wb.active
ws1.title = "CurrencyDynamic"
currency = NamedStyle(name="currency")
currency.number_format = "Currency [4]"
wb.add_named_style(currency)

for i in range(0, len(D[0])):
    ws1.append([D[0][i], D[1][i], D[2][i], E[0][i], E[1][i], E[2][i]])

for i in range(2, len(D[0])+1):
    for k in range(1, 6):
        if ws1.cell(column=k, row=i).value[0] == "'":
            ws1.cell(column=k, row=i).value = ws1.cell(column=k, row=i).value[1:]
    ws1.cell(column=7, row=i, value=("=B"+str(i)+"/E"+str(i)))


for c in range(1, 10):
    for r in range(1, len(D[0])+1):
        # ws1.cell(column=c, row=r).font = Font(bold=True)
        if c in [2, 3, 5, 6, 7]:
            ws1.cell(column=c, row=r).style = currency
            ws1.cell(column=c, row=r).style = "currency"


wb.save(filename=RESULTFILE)



# Отправляем SMTP-письмо -----------------------------------------------
def sendmessage(SMTP_SERVER, SMTP_PORT, SENDER_LOGIN, SENDER_PASSWORD, RECIEVER_EMAIL, SUBJECT, BEGIN_MESSAGE):
    if (len(D[0]) % 10) == 1 and (len(D[0]) % 100) != 11:
        finalofmessage = " строку."
    elif (len(D[0]) % 10) < 5 and (len(D[0]) % 10) > 1 and ((len(D[0]) % 100) >14 or (len(D[0]) % 100) < 11):
        finalofmessage = " строки."
    else:
        finalofmessage = " строк."
    MESSAGE_BODY = BEGIN_MESSAGE + str(len(D[0])) + finalofmessage
    MESSAGE = EmailMessage()
    MESSAGE['Subject'] = SUBJECT
    MESSAGE['From'] = SENDER_LOGIN
    MESSAGE['To'] = RECIEVER_EMAIL
    MESSAGE.set_content(MESSAGE_BODY)
    with open(RESULTFILE, 'rb') as f:
        file_data = f.read()
    MESSAGE.add_attachment(file_data, maintype = 'file', subtype='xlsx', filename = RESULTFILE)
    with smtplib.SMTP_SSL (SMTP_SERVER, SMTP_PORT) as smtp:
        # smtp.ehlo()
        # smtp.starttls()
        # smtp.ehlo()
        smtp.login(SENDER_LOGIN, SENDER_PASSWORD, initial_response_ok=True)
        smtp.send_message(MESSAGE)


sendmessage(SMTP_SERVER, SMTP_PORT, SENDER_LOGIN, SENDER_PASSWORD, RECIEVER_EMAIL, SUBJECT, BEGIN_MESSAGE)
